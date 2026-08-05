"""EU-EMA — Applications for new human medicines under evaluation（月報 xlsx）。

與其他三源本質不同：這是**目前審查中的完整清單快照**，不是當期發生的事件，
而且逐筆沒有日期。全部套檔頭的 "Data extracted"，並在報表標明性質 —— 否則
14 天視窗會把一份月快照誤讀成「這兩週有 118 件新申請」。

回報是它欄位最豐富：INN、適應症、孤兒藥、PRIME、加速審查，全都白拿。
"""

import io
import re
import warnings
from datetime import date

import openpyxl
from bs4 import BeautifulSoup

# EMA 的 xlsx 定義了一個 openpyxl 不接受的列印範圍。與資料無關，每次跑都吵一次。
warnings.filterwarnings("ignore", message="Print area cannot be set", module="openpyxl")

from ..http import Fetcher, FetchError
from ..model import Article, parse_date

SOURCE = "EU-EMA"
INDEX_URL = "https://www.ema.europa.eu/en/medicines/medicines-human-use-under-evaluation"
BASE = "https://www.ema.europa.eu"

_MONTHLY = re.compile(
    r"applications-new-human-medicines-under-evaluation-([a-z]+)-(\d{4})_en\.xlsx")
_MONTHS = ["january", "february", "march", "april", "may", "june",
           "july", "august", "september", "october", "november", "december"]

_EXTRACTED = re.compile(r"Data extracted\s+(.+)", re.I)


def fetch(fetcher: Fetcher, base_date: date) -> list[Article]:
    xlsx_url = _latest_monthly_url(fetcher)
    workbook = openpyxl.load_workbook(io.BytesIO(fetcher.get_bytes(xlsx_url)),
                                      read_only=True, data_only=True)
    rows = list(workbook[workbook.sheetnames[0]].iter_rows(values_only=True))

    header_idx, columns = _locate_header(rows)
    extracted = _extracted_date(rows[:header_idx])

    articles = []
    for row in rows[header_idx + 1:]:
        inn = _cell(row, columns.get("inn"))
        if not inn:
            continue
        articles.append(Article.make(
            source=SOURCE,
            title=f"EMA 審查中：{inn}",
            url=xlsx_url,
            pub_date=extracted,
            ingredient=inn,
            summary=_cell(row, columns.get("indication")) or "",
            extras={
                "kind": "under_evaluation",
                "snapshot": True,   # 報表據此標示「審查中清單」而非當期事件
                "orphan": _yes(_cell(row, columns.get("orphan"))),
                "prime": _yes(_cell(row, columns.get("prime"))),
                "accelerated": _yes(_cell(row, columns.get("accelerated"))),
                "ema_product_no": _cell(row, columns.get("product_no")) or "",
            },
        ))

    if not articles:
        raise FetchError(f"{xlsx_url} 解析出 0 筆 —— EMA 版型可能變了")
    return articles


def _latest_monthly_url(fetcher: Fetcher) -> str:
    """索引頁掛著數十份月報，依檔名的年月挑最新一份（不要相信頁面上的排列順序）。"""
    soup = BeautifulSoup(fetcher.get_text(INDEX_URL), "html.parser")
    best: tuple[tuple[int, int], str] | None = None
    for a in soup.find_all("a", href=True):
        m = _MONTHLY.search(a["href"])
        if not m or m.group(1) not in _MONTHS:
            continue
        key = (int(m.group(2)), _MONTHS.index(m.group(1)) + 1)
        if best is None or key > best[0]:
            href = a["href"]
            best = (key, href if href.startswith("http") else BASE + href)

    if best is None:
        raise FetchError("EMA 索引頁找不到任何月報 xlsx —— 版型可能變了")
    return best[1]


def _locate_header(rows) -> tuple[int, dict[str, int]]:
    """表頭不在第 1 列（實測在第 15 列），而且會漂 —— 從欄名字串定位，不寫死索引。"""
    wanted = {
        "inn": "international non-proprietary name",
        "indication": "indication",
        "product_no": "prod. number",
        "orphan": "orphan",
        "prime": "is prime",
        "accelerated": "accelerated assessment",
    }
    for idx, row in enumerate(rows[:60]):
        cells = [str(c).lower().strip() if c is not None else "" for c in row]
        # 必須「以欄名開頭」而非只是包含 —— 表頭上方有一段說明文字，
        # 裡面也提到 international non-proprietary name，只用 `in` 會match到那一列，
        # 於是真表頭被當成資料收進去（實測會多出 1 筆假成分）。
        if not any(c.startswith(wanted["inn"]) for c in cells):
            continue
        columns = {}
        for key, needle in wanted.items():
            for col, text in enumerate(cells):
                if needle in text:
                    columns[key] = col
                    break
        return idx, columns

    raise FetchError("EMA xlsx 找不到 INN 表頭列 —— 版型可能變了")


def _extracted_date(header_rows) -> date | None:
    for row in header_rows:
        for cell in row:
            if cell is None:
                continue
            if isinstance(cell, date):
                return cell if not hasattr(cell, "date") else cell.date()
            m = _EXTRACTED.search(str(cell))
            if m:
                return parse_date(m.group(1).strip()) or _loose_date(m.group(1))
    return None


def _loose_date(text: str) -> date | None:
    from datetime import datetime
    for fmt in ("%d %B %Y", "%d %b %Y"):
        try:
            return datetime.strptime(text.strip(), fmt).date()
        except ValueError:
            continue
    return None


def _cell(row, idx):
    if idx is None or idx >= len(row) or row[idx] is None:
        return None
    return str(row[idx]).strip() or None


def _yes(value) -> bool:
    return str(value).strip().upper() == "Y" if value else False
