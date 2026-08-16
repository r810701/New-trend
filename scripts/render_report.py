from __future__ import annotations

#!/usr/bin/env python3
"""data/raw_<date>.csv → reports/report_<date>.{html,xlsx}

四大區域設計：
1. 【區域一】藥物基本資訊 (成分粗體、英文名、適應症/標靶分類、原廠)
2. 【區域二】實證文獻 (PubMed 近一年第三期試驗論文量，隨論文越多點越大顆)
3. 【區域三】藥政進度色溫條 (取消申請留審查：實證 ➔ 審查 ➔ 核准，實證連結與 PubMed 一致)
4. 【區域四】新聞聲量 (真實報導；沒檢出就純文字顯示『尚未檢出即時新聞』)
"""

import argparse
import csv
import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from jinja2 import Template
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from newtrend import DATA_DIR, REPORTS_DIR
from newtrend.aggregate import (
    _build_category_str,
    _build_news_info,
    _eval_evidence_info,
    _eval_foreign_stages,
    _eval_gap_status,
    _eval_taiwan_stages,
    _rollup_flags,
)
from newtrend.model import Article

_RAW_NAME = re.compile(r"^raw_(\d{4}-\d{2}-\d{2})\.csv$")

HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="zh-TW">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>跨國新藥審查進度與台灣落差對比儀表板 - {{ target_date }}</title>
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&family=Noto+Sans+TC:wght@400;500;700;900&family=JetBrains+Mono:wght@500;700&display=swap" rel="stylesheet">
    <style>
        :root {
            --bg-body: #f8fafc;
            --bg-card: #ffffff;
            --border-color: #e2e8f0;
            --border-hover: #cbd5e1;
            --text-main: #0f172a;
            --text-secondary: #475569;
            --text-muted: #64748b;
            
            /* 色溫階梯調色 */
            --temp-blue-grad: linear-gradient(135deg, #2563eb 0%, #1d4ed8 100%);
            --temp-teal-grad: linear-gradient(135deg, #059669 0%, #047857 100%);
            --temp-amber-grad: linear-gradient(135deg, #d97706 0%, #b45309 100%);
            --temp-red-grad: linear-gradient(135deg, #dc2626 0%, #b91c1c 100%);
            --temp-inactive-bg: #f1f5f9;
            --temp-inactive-text: #94a3b8;

            --gap-danger-bg: #fef2f2;
            --gap-danger-text: #b91c1c;
            --gap-danger-border: #fca5a5;

            --gap-warning-bg: #fffbeb;
            --gap-warning-text: #b45309;
            --gap-warning-border: #fde68a;

            --gap-success-bg: #f0fdf4;
            --gap-success-text: #15803d;
            --gap-success-border: #86efac;

            --gap-primary-bg: #eff6ff;
            --gap-primary-text: #1d4ed8;
            --gap-primary-border: #bfdbfe;
        }

        * { box-sizing: border-box; margin: 0; padding: 0; }
        body {
            background-color: var(--bg-body);
            color: var(--text-main);
            font-family: 'Inter', 'Noto Sans TC', -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
            line-height: 1.5;
            -webkit-font-smoothing: antialiased;
        }

        /* 頂部 Header (簡潔沉穩) */
        .app-header {
            background: linear-gradient(135deg, #0b1329 0%, #1e293b 100%);
            color: #ffffff;
            padding: 2.2rem 2.5rem 2rem;
            box-shadow: 0 4px 20px rgba(11, 19, 41, 0.2);
        }
        .header-container {
            max-width: 1720px;
            margin: 0 auto;
            display: flex;
            justify-content: space-between;
            align-items: center;
            flex-wrap: wrap;
            gap: 1.5rem;
        }
        .header-title-group h1 {
            font-size: 1.75rem;
            font-weight: 800;
            letter-spacing: -0.02em;
            display: flex;
            align-items: center;
            gap: 0.6rem;
        }
        .header-title-group p {
            color: #94a3b8;
            font-size: 0.95rem;
            margin-top: 0.35rem;
        }
        .header-meta {
            display: flex;
            align-items: center;
            gap: 1rem;
            font-size: 0.82rem;
            color: #cbd5e1;
            margin-top: 0.75rem;
            flex-wrap: wrap;
        }
        .meta-tag {
            background: rgba(255, 255, 255, 0.1);
            padding: 3px 10px;
            border-radius: 999px;
            border: 1px solid rgba(255, 255, 255, 0.15);
        }
        .meta-tag.ok { border-color: #22c55e; color: #4ade80; }
        .meta-tag.fail { border-color: #ef4444; color: #f87171; }
        .btn-header {
            background: #ffffff;
            color: #0f172a;
            font-weight: 700;
            font-size: 0.85rem;
            padding: 9px 18px;
            border-radius: 8px;
            border: none;
            cursor: pointer;
            text-decoration: none;
            display: inline-flex;
            align-items: center;
            gap: 6px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.15);
            transition: all 0.2s;
        }
        .btn-header:hover {
            background: #f8fafc;
            transform: translateY(-1px);
        }

        /* 主要內容區 */
        .main-container {
            max-width: 1720px;
            margin: 2rem auto 3.5rem;
            padding: 0 1.5rem;
        }

        /* 表頭 (從左到右四大區域) */
        .table-header-bar {
            display: grid;
            grid-template-columns: 24% 24% 28% 24%;
            padding: 13px 20px;
            background: #f1f5f9;
            border-radius: 10px 10px 0 0;
            border: 1px solid var(--border-color);
            border-bottom: none;
            font-size: 0.85rem;
            font-weight: 800;
            color: #334155;
            letter-spacing: 0.02em;
        }

        /* 藥物卡片列表 */
        .drug-list {
            display: flex;
            flex-direction: column;
            gap: 0.85rem;
        }
        .drug-card {
            background: var(--bg-card);
            border: 1px solid var(--border-color);
            border-radius: 12px;
            box-shadow: 0 2px 6px rgba(15, 23, 42, 0.03);
            transition: border-color 0.2s, box-shadow 0.2s;
            overflow: hidden;
        }
        .drug-card:hover {
            border-color: var(--border-hover);
            box-shadow: 0 6px 18px rgba(15, 23, 42, 0.06);
        }

        .card-row {
            display: grid;
            grid-template-columns: 24% 24% 28% 24%;
            padding: 1.4rem 1.6rem;
            align-items: center;
            gap: 1.5rem;
        }

        /* 【區域一】藥物基本資訊 */
        .col-basic {
            display: flex;
            flex-direction: column;
            justify-content: center;
            padding-right: 0.8rem;
            border-right: 1px solid #f1f5f9;
        }
        .drug-name {
            font-size: 1.3rem;
            font-weight: 800;
            color: var(--text-main);
            letter-spacing: -0.015em;
            line-height: 1.2;
        }
        .drug-subname {
            font-size: 0.92rem;
            color: #475569;
            font-weight: 500;
            margin-top: 2px;
        }
        .drug-category {
            font-size: 0.88rem;
            color: #64748b;
            font-weight: 500;
            margin-top: 4px;
        }
        .drug-company-tag {
            font-size: 0.76rem;
            color: #94a3b8;
            margin-top: 4px;
            font-weight: 500;
        }

        /* 【區域二】實證文獻 (論文越多點越大顆) */
        .col-evidence {
            display: flex;
            flex-direction: column;
            gap: 0.6rem;
            padding-right: 0.8rem;
            border-right: 1px solid #f1f5f9;
        }
        .evidence-header {
            font-size: 0.86rem;
            font-weight: 700;
            color: #64748b;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }
        .evidence-display-box {
            background: #f8fafc;
            border: 1px solid #e2e8f0;
            border-radius: 8px;
            padding: 9px 12px;
            display: flex;
            align-items: center;
            gap: 12px;
            transition: all 0.2s;
        }
        .evidence-display-box:hover {
            background: #ffffff;
            border-color: #cbd5e1;
            box-shadow: 0 2px 8px rgba(15, 23, 42, 0.04);
        }
        
        /* 實證文獻圓點 (隨論文篇數動態變大) */
        .ev-bubble-wrap {
            width: 44px;
            height: 44px;
            display: flex;
            justify-content: center;
            align-items: center;
            flex-shrink: 0;
        }
        .ev-bubble {
            border-radius: 50%;
            display: flex;
            justify-content: center;
            align-items: center;
            color: #ffffff;
            font-weight: 800;
            font-family: 'JetBrains Mono', monospace;
            transition: transform 0.2s, box-shadow 0.2s;
            cursor: pointer;
            box-shadow: 0 2px 6px rgba(37, 99, 235, 0.25);
            text-decoration: none;
        }
        .ev-bubble:hover {
            transform: scale(1.15);
            box-shadow: 0 4px 12px rgba(37, 99, 235, 0.4);
        }
        .dot-ev-none {
            background: #cbd5e1;
            color: #64748b;
            font-size: 0.65rem;
            box-shadow: none;
        }
        .dot-ev-sm {
            background: linear-gradient(135deg, #60a5fa 0%, #3b82f6 100%);
            font-size: 0.72rem;
        }
        .dot-ev-md {
            background: linear-gradient(135deg, #3b82f6 0%, #2563eb 100%);
            font-size: 0.8rem;
        }
        .dot-ev-lg {
            background: linear-gradient(135deg, #2563eb 0%, #1d4ed8 100%);
            font-size: 0.88rem;
        }
        .dot-ev-xl {
            background: linear-gradient(135deg, #1d4ed8 0%, #1e3a8a 100%);
            font-size: 0.95rem;
            border: 2px solid #bfdbfe;
        }

        .ev-meta-text {
            display: flex;
            flex-direction: column;
            gap: 2px;
            flex: 1;
            min-width: 0;
        }
        .ev-title-line {
            font-size: 0.82rem;
            font-weight: 800;
            color: var(--text-main);
            display: flex;
            justify-content: space-between;
            align-items: center;
        }
        .ev-desc-line {
            font-size: 0.74rem;
            color: var(--text-secondary);
            line-height: 1.25;
        }
        .ev-link-btn {
            font-size: 0.72rem;
            font-weight: 700;
            color: #2563eb;
            text-decoration: none;
            display: inline-flex;
            align-items: center;
            gap: 3px;
            margin-top: 3px;
        }
        .ev-link-btn:hover { text-decoration: underline; }

        /* 【區域三】藥政階段 (取消申請留審查：實證➔審查➔核准，實證連結與 PubMed 一致) */
        .col-regulatory {
            display: flex;
            flex-direction: column;
            gap: 0.6rem;
            padding-right: 0.8rem;
            border-right: 1px solid #f1f5f9;
        }
        .reg-header {
            font-size: 0.86rem;
            font-weight: 700;
            color: #64748b;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }

        /* 色溫條容器 */
        .temp-bar-mini {
            display: flex;
            width: 100%;
            height: 32px;
            border-radius: 6px;
            overflow: hidden;
            border: 1px solid #cbd5e1;
            background: #ffffff;
            margin-bottom: 2px;
        }
        .temp-seg {
            flex: 1;
            display: flex;
            flex-direction: column;
            justify-content: center;
            align-items: center;
            text-align: center;
            text-decoration: none;
            padding: 1px 3px;
            border-right: 1px solid rgba(255, 255, 255, 0.4);
            transition: all 0.2s;
            color: inherit;
        }
        .temp-seg:last-child { border-right: none; }
        .seg-name { font-size: 0.72rem; font-weight: 700; line-height: 1.1; white-space: nowrap; }

        .seg-active-f1 { background: var(--temp-blue-grad); color: #ffffff; }
        .seg-active-f2 { background: var(--temp-amber-grad); color: #ffffff; }
        .seg-active-f3 { background: var(--temp-red-grad); color: #ffffff; }

        .seg-active-t1 { background: var(--temp-blue-grad); color: #ffffff; }
        .seg-active-t2 { background: var(--temp-teal-grad); color: #ffffff; }
        .seg-active-t3 { background: var(--temp-red-grad); color: #ffffff; }

        .seg-inactive { background: var(--temp-inactive-bg); color: var(--temp-inactive-text); border-right: 1px solid #e2e8f0; }
        .seg-inactive:hover { background: #e2e8f0; }

        .gap-badge-sm {
            font-size: 0.72rem;
            font-weight: 800;
            padding: 2px 7px;
            border-radius: 4px;
            letter-spacing: 0.01em;
            margin-bottom: 4px;
            display: inline-block;
        }
        .badge-gap-danger { background: var(--gap-danger-bg); color: var(--gap-danger-text); border: 1px solid var(--gap-danger-border); }
        .badge-gap-warning { background: var(--gap-warning-bg); color: var(--gap-warning-text); border: 1px solid var(--gap-warning-border); }
        .badge-gap-success { background: var(--gap-success-bg); color: var(--gap-success-text); border: 1px solid var(--gap-success-border); }
        .badge-gap-primary { background: var(--gap-primary-bg); color: var(--gap-primary-text); border: 1px solid var(--gap-primary-border); }
        .badge-gap-tw { background: #f1f5f9; color: #334155; border: 1px solid #cbd5e1; }
        .badge-gap-secondary { background: #f1f5f9; color: #64748b; border: 1px solid #e2e8f0; }

        /* 【區域四】新聞聲量 (真實報導；沒檢出就純文字顯示尚未檢出) */
        .col-news {
            display: flex;
            flex-direction: column;
            gap: 0.55rem;
        }
        .news-header-label {
            font-size: 0.86rem;
            font-weight: 700;
            color: #64748b;
            display: flex;
            align-items: center;
            gap: 4px;
        }
        .news-card-real {
            background: #f8fafc;
            border: 1px solid #e2e8f0;
            border-left: 3.5px solid #2563eb;
            border-radius: 6px;
            padding: 7px 10px;
            font-size: 0.76rem;
            transition: all 0.15s;
        }
        .news-card-real:hover {
            background: #ffffff;
            border-color: var(--border-hover);
            box-shadow: 0 2px 8px rgba(15, 23, 42, 0.05);
        }
        .news-mini-top {
            display: flex;
            justify-content: space-between;
            align-items: center;
            gap: 4px;
            margin-bottom: 2px;
        }
        .news-source-tag {
            font-size: 0.68rem;
            font-weight: 800;
            padding: 1px 6px;
            border-radius: 3px;
            background: #1e293b;
            color: #ffffff;
        }
        .news-mini-title {
            font-weight: 700;
            color: var(--text-main);
            line-height: 1.3;
            display: -webkit-box;
            -webkit-line-clamp: 1;
            -webkit-box-orient: vertical;
            overflow: hidden;
        }
        .verify-btn {
            display: inline-flex;
            align-items: center;
            gap: 3px;
            font-size: 0.7rem;
            font-weight: 700;
            color: #2563eb;
            text-decoration: none;
            padding: 2px 7px;
            background: #ffffff;
            border: 1px solid #bfdbfe;
            border-radius: 4px;
            transition: all 0.15s;
        }
        .verify-btn:hover {
            background: #2563eb;
            color: #ffffff;
        }

        /* 尚未檢出即時新聞之純粹狀態 */
        .news-empty-simple {
            background: #f8fafc;
            border: 1px dashed #cbd5e1;
            border-radius: 6px;
            padding: 12px 14px;
            font-size: 0.78rem;
            font-weight: 600;
            color: #94a3b8;
            text-align: center;
        }

        /* 頁尾 */
        .app-footer {
            text-align: center;
            color: var(--text-muted);
            font-size: 0.82rem;
            padding: 2.5rem 0 3.5rem;
            border-top: 1px solid var(--border-color);
            margin-top: 3rem;
        }

        /* 響應式佈局 */
        @media (max-width: 1300px) {
            .table-header-bar { display: none; }
            .card-row { grid-template-columns: 1fr 1fr; gap: 1.4rem; }
            .col-basic, .col-evidence, .col-regulatory { border-right: none; padding-right: 0; }
        }
        @media (max-width: 768px) {
            .card-row { grid-template-columns: 1fr; }
        }
    </style>
</head>
<body>

    <!-- 頂部 Header -->
    <header class="app-header">
        <div class="header-container">
            <div class="header-title-group">
                <h1>🧬 跨國新藥審查進度與台灣落差對比儀表板</h1>
                <p>US-FDA／EU-EMA 官方審查清單 ｜ PubMed 近一年三期實證文獻 ｜ TW-TFDA/健保署 台灣藥政進度</p>
                <div class="header-meta">
                    <span>📅 基準日期：<strong>{{ target_date }}</strong></span>
                    <span>📦 追蹤標的：<strong>{{ total_items }}</strong> 項</span>
                    <span>｜ 來源狀態：
                        {% for s in source_status %}
                            <span class="meta-tag {{ 'ok' if s.ok else 'fail' }}">{{ s.name }} {{ '✓' if s.ok else '✗' }}</span>
                        {% endfor %}
                    </span>
                </div>
            </div>
            <div class="header-actions">
                <button onclick="window.print()" class="btn-header">🖨️ 列印 / 另存 PDF</button>
            </div>
        </div>
    </header>

    <main class="main-container">

        <!-- 表頭 (從左到右四大區域) -->
        <div class="table-header-bar">
            <div>【區域一】藥物基本資訊</div>
            <div>【區域二】實證文獻 (近一年 Phase 3 論文)</div>
            <div>【區域三】藥政階段 (國外 vs 台灣色溫條)</div>
            <div>【區域四】新聞聲量 (真實報導)</div>
        </div>

        <!-- 藥物列表清單 -->
        <div class="drug-list" id="drugList">
            {% for item in results %}
            <article class="drug-card">
                <div class="card-row">
                    
                    <!-- 【區域一】藥物基本資訊 -->
                    <div class="col-basic">
                        <div class="drug-name">{{ item.drug_ingredient }}</div>
                        <div class="drug-subname">
                            {% if item.brand %}
                                ({{ item.brand }})
                            {% else %}
                                ({{ item.drug_ingredient|lower }})
                            {% endif %}
                        </div>
                        <div class="drug-category">
                            {{ item.category or '新藥標靶 / 臨床評估' }}
                        </div>
                        {% if item.company %}
                            <div class="drug-company-tag">🏢 {{ item.company }}</div>
                        {% endif %}
                    </div>

                    <!-- 【區域二】實證文獻 (PubMed 近一年三期論文，點越大論文越多) -->
                    <div class="col-evidence">
                        <div class="evidence-header">
                            <span>實證文獻</span>
                            <span style="font-size:0.74rem; color:#2563eb; font-weight:700;">{{ item.evidence.phase }}</span>
                        </div>
                        <div class="evidence-display-box">
                            <div class="ev-bubble-wrap">
                                <a href="{{ item.evidence.url }}" target="_blank" rel="noopener noreferrer" 
                                   class="ev-bubble {{ item.evidence.dot_class }}" 
                                   style="width: {{ item.evidence.dot_size }}px; height: {{ item.evidence.dot_size }}px;"
                                   title="點擊查驗 PubMed 近一年 Phase 3 文獻">
                                    {{ item.evidence.count }}
                                </a>
                            </div>
                            <div class="ev-meta-text">
                                <div class="ev-title-line">
                                    <span>{{ item.evidence.count }} 篇文獻</span>
                                </div>
                                <div class="ev-desc-line">{{ item.evidence.desc }}</div>
                                <a href="{{ item.evidence.url }}" target="_blank" rel="noopener noreferrer" class="ev-link-btn">
                                    PubMed 檢索 ↗
                                </a>
                            </div>
                        </div>
                    </div>

                    <!-- 【區域三】藥政階段 (取消申請留審查：實證➔審查➔核准，實證連結與 PubMed 一致) -->
                    <div class="col-regulatory">
                        <div class="reg-header">
                            <span>藥政進度色溫條</span>
                            <span class="gap-badge-sm {{ item.gap.badge_class }}">{{ item.gap.label }}</span>
                        </div>

                        <!-- 國外 3 階段色溫條 (實證 ➔ 審查 ➔ 核准) -->
                        <div class="temp-bar-mini" title="國外 3 階段：實證 ➔ 審查 ➔ 核准">
                            {% set fs1 = item.foreign_stages.stages[0] %}
                            <a href="{{ fs1.url }}" target="_blank" rel="noopener noreferrer" 
                               class="temp-seg {{ 'seg-active-f1' if fs1.reached else 'seg-inactive' }}" title="國外實證：{{ fs1.title }} (點擊連至 PubMed)">
                                <span class="seg-name">實證</span>
                            </a>
                            {% set fs2 = item.foreign_stages.stages[1] %}
                            <a href="{{ fs2.url }}" target="_blank" rel="noopener noreferrer" 
                               class="temp-seg {{ 'seg-active-f2' if fs2.reached else 'seg-inactive' }}" title="國外審查：{{ fs2.title }} (點擊查驗 EMA/FDA)">
                                <span class="seg-name">國外審查</span>
                            </a>
                            {% set fs3 = item.foreign_stages.stages[2] %}
                            <a href="{{ fs3.url }}" target="_blank" rel="noopener noreferrer" 
                               class="temp-seg {{ 'seg-active-f3' if fs3.reached else 'seg-inactive' }}" title="國外核准：{{ fs3.title }} (點擊查驗 FDA Novel Approvals)">
                                <span class="seg-name">國外核准</span>
                            </a>
                        </div>

                        <!-- 台灣 3 階段色溫條 (審查 ➔ 獲藥證 ➔ 健保給付) -->
                        <div class="temp-bar-mini" title="台灣 3 階段：台灣審查 ➔ 取得藥證 ➔ 核准健保給付">
                            {% set ts1 = item.taiwan_stages.stages[0] %}
                            <a href="{{ ts1.url }}" target="_blank" rel="noopener noreferrer" 
                               class="temp-seg {{ 'seg-active-t1' if ts1.reached else 'seg-inactive' }}" title="台灣審查：{{ ts1.title }}">
                                <span class="seg-name">台灣審查</span>
                            </a>
                            {% set ts2 = item.taiwan_stages.stages[1] %}
                            <a href="{{ ts2.url }}" target="_blank" rel="noopener noreferrer" 
                               class="temp-seg {{ 'seg-active-t2' if ts2.reached else 'seg-inactive' }}" title="取得藥證：{{ ts2.title }} (點擊查驗 TFDA 審查報告 PDF)">
                                <span class="seg-name">取得藥證</span>
                            </a>
                            {% set ts3 = item.taiwan_stages.stages[2] %}
                            <a href="{{ ts3.url }}" target="_blank" rel="noopener noreferrer" 
                               class="temp-seg {{ 'seg-active-t3' if ts3.reached else 'seg-inactive' }}" title="健保給付：{{ ts3.title }}">
                                <span class="seg-name">健保給付</span>
                            </a>
                        </div>
                    </div>

                    <!-- 【區域四】新聞聲量 (真實報導；沒檢出就純文字顯示尚未檢出) -->
                    <div class="col-news">
                        {% if item.news_info.has_news %}
                            {% for news in item.news_info.articles %}
                            <div class="news-card-real">
                                <div class="news-mini-top">
                                    <span class="news-source-tag">{{ news.source }}</span>
                                    <a href="{{ news.url }}" target="_blank" rel="noopener noreferrer" class="verify-btn">認證 ↗</a>
                                </div>
                                <div class="news-mini-title" title="{{ news.title }}">{{ news.title }}</div>
                            </div>
                            {% endfor %}
                        {% else %}
                            <div class="news-empty-simple">尚未檢出即時新聞</div>
                        {% endif %}
                    </div>

                </div>
            </article>
            {% endfor %}
        </div>

    </main>

    <footer class="app-footer">
        <p>跨國新藥審查進度與台灣落差對比報告 ｜ 產生時間：{{ generated_at }}</p>
        <p style="margin-top: 4px; font-size: 0.75rem;">資料來源：US-FDA, EU-EMA, TW-TFDA, CDE, TW-NHI, PubMed (近一年三期文獻), ClinicalTrials.gov, Fierce Biotech, 環球生技</p>
    </footer>
</body>
</html>
"""


def parse_args(argv=None):
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--input", type=Path,
                   help="指定 data/raw_<date>.csv（預設取日期最新的一份）")
    p.add_argument("--target-date", default=os.environ.get("TARGET_DATE"),
                   help="改讀該日期的 CSV（也吃 TARGET_DATE 環境變數）")
    return p.parse_args(argv)


def main(argv=None) -> int:
    args = parse_args(argv)

    if args.input:
        csv_path = args.input
    elif args.target_date:
        csv_path = DATA_DIR / f"raw_{args.target_date.strip()}.csv"
    else:
        csv_path = _latest_csv()

    if csv_path is None or not csv_path.exists():
        print(f"找不到 {csv_path or 'data/raw_*.csv'} —— 先跑 scripts/fetch.py",
              file=sys.stderr)
        return 1

    target_date = _date_from_name(csv_path)
    results = _read_rows(csv_path)
    meta = _read_meta(csv_path)

    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    _write_xlsx(REPORTS_DIR / f"report_{target_date}.xlsx", results)
    _write_html(REPORTS_DIR / f"report_{target_date}.html", results, target_date, meta)

    print(f"✓ reports/report_{target_date}.html / .xlsx（{len(results)} 個標的）")
    return 0


def _latest_csv() -> Path | None:
    dated = []
    for path in DATA_DIR.glob("raw_*.csv"):
        m = _RAW_NAME.match(path.name)
        if m:
            dated.append((m.group(1), path))
    return max(dated)[1] if dated else None


def _date_from_name(path: Path) -> str:
    m = _RAW_NAME.match(path.name)
    return m.group(1) if m else datetime.now().strftime("%Y-%m-%d")


def _read_rows(path: Path) -> list[dict]:
    rows = []
    with path.open(encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            articles = json.loads(row.get("articles_json") or "[]")
            row["articles"] = articles

            flags_text = row.get("flags", "")
            phase_match = re.search(r"[一二三四1234]期", flags_text)
            phase_val = phase_match.group(0) if phase_match else None
            is_orphan = "孤兒藥" in flags_text
            is_prime = "PRIME" in flags_text
            is_accelerated = "加速審查" in flags_text

            article_objs = []
            for a in articles:
                extras = a.get("extras", {}) or {}
                if not extras.get("phase") and phase_val:
                    extras["phase"] = phase_val
                if not extras.get("orphan") and is_orphan:
                    extras["orphan"] = True
                if not extras.get("prime") and is_prime:
                    extras["prime"] = True
                if not extras.get("accelerated") and is_accelerated:
                    extras["accelerated"] = True

                source = a.get("source", "")
                title = a.get("title", "")
                if source == "US-FDA" or "fda" in source.lower() or "fda 核准" in title.lower():
                    extras["kind"] = "approval"
                elif source == "EU-EMA" or "ema" in source.lower() or "ema 審查" in title.lower():
                    extras["kind"] = "under_evaluation"
                elif source == "TW-TFDA" or "tfda" in source.lower() or "tfda 核准" in title.lower():
                    extras["kind"] = "approval"

                article_objs.append(
                    Article.make(
                        source=source,
                        title=title,
                        url=a.get("url", ""),
                        ingredient=row.get("drug_ingredient"),
                        company=row.get("company"),
                        summary=a.get("summary", ""),
                        extras=extras,
                    )
                )

            name = row.get("drug_ingredient", "")
            has_inn = row.get("has_inn") == "True" or row.get("has_inn") is True
            evidence_info = _eval_evidence_info(name, article_objs, flags_text)
            f_stages = _eval_foreign_stages(name, article_objs, evidence_info["url"])
            t_stages = _eval_taiwan_stages(name, article_objs)
            gap = _eval_gap_status(f_stages, t_stages, has_inn)
            news_info = _build_news_info(article_objs)

            row["evidence"] = evidence_info
            row["foreign_stages"] = f_stages
            row["taiwan_stages"] = t_stages
            row["gap"] = gap
            row["news_info"] = news_info
            row["indication"] = row.get("indication") or (articles[0].get("summary", "") if articles else "")
            row["category"] = _build_category_str(row["indication"], _rollup_flags(article_objs))
            row["brand"] = row.get("brand", "")

            rows.append(row)

    # 排序：落差嚴重程度優先（國外已核准 ➔ 國外審查中 ➔ 台灣核准 ➔ 其他），次依文獻篇數
    gap_order = {
        "foreign_approved_tw_pending": 1,
        "foreign_review_tw_pending": 2,
        "tw_approved_nhi_pending": 3,
        "tw_nhi_reimbursed": 4,
        "tw_exclusive": 5,
        "early_stage": 6,
    }
    rows.sort(key=lambda r: (
        gap_order.get(r["gap"]["code"], 99),
        -(r["evidence"]["count"]),
        r.get("latest_date") or "0000-00-00",
    ), reverse=False)

    return rows


def _read_meta(csv_path: Path) -> dict:
    meta_path = csv_path.with_name(csv_path.stem + ".meta.json")
    if not meta_path.exists():
        return {}
    return json.loads(meta_path.read_text(encoding="utf-8"))


def _write_html(path: Path, results: list[dict], target_date: str, meta: dict) -> None:
    sources = meta.get("sources") or {}
    filtered_sources = {k: v for k, v in sources.items() if k != "nice"}

    html = Template(HTML_TEMPLATE).render(
        results=results,
        total_items=len(results),
        target_date=target_date,
        source_status=[
            {
                "name": n,
                "ok": s.get("ok", False),
                "reason_label": s.get("reason_label", "未知錯誤"),
                "status_code": s.get("status_code"),
            }
            for n, s in filtered_sources.items()
        ],
        generated_at=meta.get("generated_at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    )
    path.write_text(html, encoding="utf-8")


def _write_xlsx(path: Path, results: list[dict]) -> None:
    wb = Workbook()

    sheet1 = wb.active
    sheet1.title = "新藥實證與落差對比"
    _header(sheet1, [
        "#", "藥物成分 (INN)", "商品名", "研發原廠 / 申請商", "主要適應症", "標靶 / 分類",
        "PubMed 近一年 Phase 3 論文篇數", "PubMed 檢索連結",
        "國外階段 (實證➔審查➔核准)", "國外認證連結",
        "台灣階段 (審查➔藥證➔健保)", "台灣許可證字號", "台灣審查 PDF 連結",
        "⚡ 國外 vs 台灣進度落差", "最新日期"
    ])

    for i, row in enumerate(results, 1):
        f_stages = row.get("foreign_stages", {})
        t_stages = row.get("taiwan_stages", {})
        gap = row.get("gap", {})
        evidence = row.get("evidence", {})

        f_str = " ➔ ".join([s["label"] if s["reached"] else f"({s['label']}:未出現)" for s in f_stages.get("stages", [])])
        t_str = " ➔ ".join([s["label"] if s["reached"] else f"({s['label']}:未出現)" for s in t_stages.get("stages", [])])

        f_url = f_stages.get("stages", [{}])[-1].get("url") if f_stages.get("stages") else ""
        t_url = t_stages.get("stages", [{}])[1].get("url") if len(t_stages.get("stages", [])) > 1 else ""

        sheet1.append([
            i,
            row.get("drug_ingredient", ""),
            row.get("brand", ""),
            row.get("company", ""),
            row.get("indication", ""),
            row.get("category", ""),
            f"{evidence.get('count', 0)} 篇",
            evidence.get("url", ""),
            f_str,
            f_url,
            t_str,
            t_stages.get("license_no", ""),
            t_url,
            gap.get("label", ""),
            row.get("latest_date", ""),
        ])

    widths1 = [5, 26, 16, 24, 38, 22, 22, 36, 30, 32, 28, 22, 32, 28, 12]
    for col, width in enumerate(widths1, 1):
        sheet1.column_dimensions[get_column_letter(col)].width = width
    sheet1.freeze_panes = "B2"

    wb.save(path)


def _header(sheet, names: list[str]) -> None:
    sheet.append(names)
    fill = PatternFill("solid", fgColor="0F172A")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = fill
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)


if __name__ == "__main__":
    sys.exit(main())
